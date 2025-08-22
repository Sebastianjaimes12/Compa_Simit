from flask import Flask, render_template_string, request, jsonify, send_file
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import os
import time
import json
import threading
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import platform

# Configuración OPTIMIZADA para Render
def configurar_chrome_para_render():
    options = Options()
    
    if platform.system() == "Linux":
        # Configuración agresiva para evitar timeouts
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-plugins')
        options.add_argument('--disable-images')
        options.add_argument('--disable-background-timer-throttling')
        options.add_argument('--disable-backgrounding-occluded-windows')
        options.add_argument('--disable-renderer-backgrounding')
        options.add_argument('--disable-features=TranslateUI')
        options.add_argument('--disable-ipc-flooding-protection')
        options.add_argument('--window-size=1280,720')  # Tamaño más pequeño
        options.add_argument('--memory-pressure-off')
        options.add_argument('--single-process')  # Proceso único para usar menos memoria
        
        # Timeouts más agresivos
        options.add_argument('--page-load-strategy=none')
        options.add_experimental_option('useAutomationExtension', False)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        
        options.binary_location = "/usr/bin/google-chrome"
    else:
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
    
    return options

app = Flask(__name__)

# Variables globales mejoradas
progreso_actual = {
    'estado': 'idle',
    'mensaje': 'Listo para iniciar',
    'placa_actual': '',
    'total': 0,
    'procesadas': 0,
    'porcentaje': 0,
    'resultados': [],
    'archivo_excel': '',
    'inicio_proceso': None,
    'thread_id': None
}

# Timeout para limpiar procesos colgados (15 minutos)
TIMEOUT_PROCESO = 900  # 15 minutos

def limpiar_proceso_si_colgado():
    """Limpia el estado si el proceso lleva mucho tiempo"""
    global progreso_actual
    
    if (progreso_actual['estado'] == 'processing' and 
        progreso_actual['inicio_proceso'] and
        datetime.now() - progreso_actual['inicio_proceso'] > timedelta(seconds=TIMEOUT_PROCESO)):
        
        print("🧹 Limpiando proceso colgado...")
        progreso_actual.update({
            'estado': 'idle',
            'mensaje': 'Proceso anterior cancelado por timeout',
            'placa_actual': '',
            'total': 0,
            'procesadas': 0,
            'porcentaje': 0,
            'resultados': [],
            'archivo_excel': '',
            'inicio_proceso': None,
            'thread_id': None
        })

class SimitScraper:
    def __init__(self):
        self.resultados = []
        self.driver = None
        self.proceso_cancelado = False
    
    def actualizar_progreso(self, mensaje, placa_actual='', total=0, procesadas=0):
        global progreso_actual
        
        if self.proceso_cancelado:
            return
        
        porcentaje = 0
        if total > 0:
            porcentaje = round((procesadas / total) * 100, 1)
            porcentaje = min(porcentaje, 100)
        
        progreso_actual.update({
            'mensaje': mensaje,
            'placa_actual': placa_actual,
            'total': total,
            'procesadas': procesadas,
            'porcentaje': porcentaje
        })

    def esperar_simit_completamente(self, driver):
        """Espera específica para que SIMIT esté listo"""
        try:
            print("⏳ Esperando que SIMIT esté completamente cargado...")
            
            # 1. Esperar que Angular/Vue esté listo
            WebDriverWait(driver, 15).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            time.sleep(3)
            
            # 2. Esperar específicamente por el campo de búsqueda
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "txtBusqueda"))
            )
            time.sleep(2)
            
            # 3. Verificar que el campo sea interactuable
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "txtBusqueda"))
            )
            time.sleep(2)
            
            print("✅ SIMIT completamente cargado y listo")
            return True
            
        except Exception as e:
            print(f"⚠️ Timeout esperando SIMIT: {e}")
            return False

    def navegar_a_simit_con_reintentos(self, driver):
        """Navegación robusta a SIMIT con múltiples intentos"""
        urls_simit = [
            "https://www.fcm.org.co/simit/#/home-public",
            "https://www.fcm.org.co/simit/",
            "https://fcm.org.co/simit/#/home-public"
        ]
        
        for intento, url in enumerate(urls_simit):
            try:
                print(f"🌐 Intento {intento + 1}: Navegando a {url}")
                
                driver.get(url)
                
                # Esperar que SIMIT esté listo
                if self.esperar_simit_completamente(driver):
                    print(f"✅ SIMIT cargado exitosamente en intento {intento + 1}")
                    return True
                    
            except Exception as e:
                print(f"❌ Error en intento {intento + 1}: {e}")
                if intento < len(urls_simit) - 1:
                    print("🔄 Probando siguiente URL...")
                    time.sleep(2)
                    continue
        
        # Si fallan todas las URLs, intentar estrategia manual
        print("🔧 Intentando estrategia manual...")
        try:
            driver.get("https://www.fcm.org.co/")
            time.sleep(5)
            
            # Intentar hacer click en el enlace de SIMIT si existe
            try:
                enlace_simit = driver.find_element(By.LINK_TEXT, "SIMIT")
                enlace_simit.click()
                time.sleep(5)
                
                if self.esperar_simit_completamente(driver):
                    print("✅ SIMIT cargado mediante navegación manual")
                    return True
            except:
                pass
                
        except Exception as e:
            print(f"❌ Error en estrategia manual: {e}")
        
        return False

    def detectar_multas_mejorada(self, driver, placa):
        """Detección mejorada y más rápida"""
        try:
            if self.proceso_cancelado:
                return False, 0
                
            time.sleep(1)  # Reducido de 2 a 1 segundo
            
            # MÉTODO 1: Buscar tabla específica de SIMIT
            try:
                tabla_multas = driver.find_element(By.ID, "multaTable")
                tbody = tabla_multas.find_element(By.TAG_NAME, "tbody")
                filas = tbody.find_elements(By.TAG_NAME, "tr")
                
                filas_con_multas = []
                for fila in filas:
                    if self.proceso_cancelado:
                        return False, 0
                        
                    texto_fila = fila.text.strip().lower()
                    if texto_fila and not any(palabra in texto_fila for palabra in [
                        'no se encontraron', 'sin multas', 'no hay multas', 'no tiene multas'
                    ]):
                        celdas = fila.find_elements(By.TAG_NAME, "td")
                        if len(celdas) >= 6:
                            filas_con_multas.append(fila)
                
                if len(filas_con_multas) > 0:
                    print(f"✅ MULTAS DETECTADAS: {len(filas_con_multas)} multa(s)")
                    return True, len(filas_con_multas)
                else:
                    print("✅ SIN MULTAS")
                    return False, 0
                    
            except Exception as e:
                print(f"No se encontró tabla #multaTable: {e}")
            
            # MÉTODO 2: Análisis rápido de texto
            texto_pagina = driver.page_source.lower()
            
            # Mensajes de sin multas
            sin_multas_frases = [
                "no se encontraron multas",
                "sin multas registradas", 
                "no hay multas",
                "no tiene multas"
            ]
            
            for frase in sin_multas_frases:
                if frase in texto_pagina:
                    print(f"✅ SIN MULTAS - '{frase}'")
                    return False, 0
            
            # Indicadores de multas
            if any(palabra in texto_pagina for palabra in [
                "valor a pagar", "cobro coactivo", "secretaría"
            ]):
                print("✅ MULTAS DETECTADAS")
                return True, 1
            
            return False, 0
            
        except Exception as e:
            print(f"❌ Error en detección: {e}")
            return False, 0

    def extraer_detalles_multas(self, driver, placa):
        """Extracción optimizada de detalles"""
        if self.proceso_cancelado:
            return "Proceso cancelado"
            
        detalles = ""
        try:
            tabla_multas = driver.find_element(By.ID, "multaTable")
            tbody = tabla_multas.find_element(By.TAG_NAME, "tbody")
            filas = tbody.find_elements(By.TAG_NAME, "tr")
            
            multa_count = 0
            for fila in filas:
                if self.proceso_cancelado:
                    break
                    
                try:
                    texto_fila = fila.text.strip()
                    if not texto_fila or any(palabra in texto_fila.lower() for palabra in [
                        'no se encontraron', 'sin multas'
                    ]):
                        continue
                    
                    celdas = fila.find_elements(By.TAG_NAME, "td")
                    if len(celdas) >= 6:
                        multa_count += 1
                        detalles += f"=== MULTA {multa_count} ===\n"
                        
                        # Extraer datos básicos
                        for i, etiqueta in enumerate([
                            "Tipo", "Notificación", "Placa", "Secretaría", 
                            "Infracción", "Estado", "Valor", "Valor a pagar"
                        ]):
                            if i < len(celdas):
                                valor = celdas[i].text.strip()
                                if valor:
                                    detalles += f"{etiqueta}: {valor}\n"
                        
                        detalles += "\n"
                        
                except Exception as e:
                    continue
                    
        except Exception as e:
            print(f"Error extrayendo detalles: {e}")
            detalles = "No se pudieron extraer detalles específicos"
            
        return detalles.strip() if detalles.strip() else "Sin detalles disponibles"

    def tomar_captura_optimizada(self, placa, driver):
        """Captura optimizada para Render"""
        try:
            if self.proceso_cancelado:
                return "Sin captura"
                
            if not os.path.exists("capturas"):
                os.makedirs("capturas")
            
            screenshot_path = f"capturas/{placa}_{datetime.now().strftime('%H%M%S')}.png"
            
            # Captura rápida sin scroll
            driver.save_screenshot(screenshot_path)
            
            if os.path.exists(screenshot_path):
                return screenshot_path
            else:
                return "Sin captura"
                
        except Exception as e:
            print(f"Error en captura: {e}")
            return "Sin captura"

    def buscar_placas(self, placas):
        global progreso_actual
        
        try:
            # Marcar inicio del proceso
            progreso_actual['estado'] = 'processing'
            progreso_actual['inicio_proceso'] = datetime.now()
            progreso_actual['thread_id'] = threading.current_thread().ident
            
            self.actualizar_progreso("🚀 Iniciando proceso...", total=len(placas), procesadas=0)
            
            # Configurar Chrome con timeout
            service = Service()
            options = configurar_chrome_para_render()
            
            self.actualizar_progreso("🔧 Iniciando navegador...", total=len(placas), procesadas=0)
            
            try:
                self.driver = webdriver.Chrome(service=service, options=options)
                # Timeouts MUY agresivos para evitar colgados
                self.driver.set_page_load_timeout(20)  # Reducido de 30 a 20
                self.driver.implicitly_wait(5)  # Reducido de 10 a 5
                print("✅ Chrome iniciado correctamente")
            except Exception as e:
                raise Exception(f"Error iniciando Chrome: {str(e)}")
            
            if platform.system() != "Linux":
                self.driver.maximize_window()
            
            self.actualizar_progreso("🌐 Navegando a SIMIT...", total=len(placas), procesadas=0)
            
            # Navegación robusta a SIMIT
            if not self.navegar_a_simit_con_reintentos(self.driver):
                raise Exception("No se pudo cargar SIMIT después de múltiples intentos")
            
            print("🎯 SIMIT cargado correctamente - comenzando búsquedas...")
            
            # Procesar cada placa
            for idx, placa in enumerate(placas):
                if self.proceso_cancelado:
                    break
                    
                try:
                    self.actualizar_progreso(f"🔍 Procesando: {placa}", placa, len(placas), idx)
                    
                    # Cerrar popups con timeout
                    try:
                        popup = WebDriverWait(self.driver, 2).until(
                            EC.presence_of_element_located((By.CLASS_NAME, "swal2-popup"))
                        )
                        cerrar_btn = self.driver.find_element(By.CLASS_NAME, "swal2-confirm")
                        cerrar_btn.click()
                        time.sleep(1)
                    except:
                        pass

                    # Buscar campo de placa con verificación robusta
                    try:
                        print(f"🔍 Buscando campo de búsqueda para {placa}...")
                        
                        # Verificar que el campo existe y es visible
                        campo_placa = WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.ID, "txtBusqueda"))
                        )
                        
                        # Verificar que el campo es interactuable
                        WebDriverWait(self.driver, 5).until(
                            EC.element_to_be_clickable((By.ID, "txtBusqueda"))
                        )
                        
                        # Scroll al elemento si es necesario
                        self.driver.execute_script("arguments[0].scrollIntoView(true);", campo_placa)
                        time.sleep(1)
                        
                        # Hacer click para asegurar que está activo
                        campo_placa.click()
                        time.sleep(0.5)
                        
                        # Limpiar y escribir
                        campo_placa.clear()
                        time.sleep(0.5)
                        campo_placa.send_keys(placa)
                        time.sleep(1)
                        
                        # Enviar búsqueda (múltiples métodos)
                        try:
                            campo_placa.send_keys("\n")
                        except:
                            # Método alternativo: buscar botón de búsqueda
                            try:
                                boton_buscar = self.driver.find_element(By.XPATH, "//button[contains(@class, 'btn') or contains(text(), 'Buscar')]")
                                boton_buscar.click()
                            except:
                                # Método JavaScript como último recurso
                                self.driver.execute_script("arguments[0].form.submit();", campo_placa)
                        
                        print(f"✅ Búsqueda enviada para {placa}")
                        
                        # Esperar resultados
                        time.sleep(6)
                        
                    except Exception as e:
                        print(f"❌ Error buscando {placa}: {e}")
                        # Intentar recargar SIMIT si el campo no está disponible
                        try:
                            print("🔄 Recargando SIMIT...")
                            if self.navegar_a_simit_con_reintentos(self.driver):
                                print("✅ SIMIT recargado, continuando...")
                                continue
                            else:
                                raise Exception("No se pudo recargar SIMIT")
                        except:
                            print(f"❌ Error crítico con {placa}")
                            self.resultados.append((placa, "Error", "Campo no interactuable", "Sin captura", str(e)))
                            continue
                    
                    # Detectar multas
                    tiene_multas, num_multas = self.detectar_multas_mejorada(self.driver, placa)
                    
                    # Extraer detalles si hay multas
                    detalle_multas = ""
                    if tiene_multas and not self.proceso_cancelado:
                        self.actualizar_progreso(f"📋 Extrayendo detalles de {placa}...", placa, len(placas), idx)
                        detalle_multas = self.extraer_detalles_multas(self.driver, placa)
                    
                    # Tomar captura
                    if not self.proceso_cancelado:
                        screenshot_path = self.tomar_captura_optimizada(placa, self.driver)
                    else:
                        screenshot_path = "Sin captura"
                    
                    estado_multas = "Sí" if tiene_multas else "No"
                    self.resultados.append((placa, estado_multas, "Éxito", screenshot_path, detalle_multas))
                    
                    # Actualizar progreso
                    procesadas_actual = idx + 1
                    self.actualizar_progreso(f"✅ {placa}: {estado_multas} multas", placa, len(placas), procesadas_actual)
                    
                    time.sleep(1)  # Pausa reducida
                    
                except Exception as e:
                    procesadas_actual = idx + 1
                    self.actualizar_progreso(f"❌ Error en {placa}", placa, len(placas), procesadas_actual)
                    self.resultados.append((placa, "Error", "Error", "Sin captura", str(e)))
            
            if not self.proceso_cancelado:
                # Generar Excel
                self.actualizar_progreso("📊 Generando Excel...", total=len(placas), procesadas=len(placas))
                archivo_excel = self.guardar_resultados_en_excel()
                
                if archivo_excel and os.path.exists(archivo_excel):
                    progreso_actual.update({
                        'estado': 'completed',
                        'resultados': self.resultados,
                        'archivo_excel': archivo_excel,
                        'porcentaje': 100,
                        'procesadas': len(placas),
                        'total': len(placas),
                        'mensaje': '🎉 Proceso completado. Excel listo para descarga.'
                    })
                else:
                    raise Exception("Error generando Excel")
            
        except Exception as e:
            if not self.proceso_cancelado:
                progreso_actual.update({
                    'estado': 'error',
                    'mensaje': f"💥 Error: {str(e)}",
                    'porcentaje': 0
                })
                print(f"ERROR: {e}")
        finally:
            try:
                if self.driver:
                    self.driver.quit()
                    print("🔒 Navegador cerrado")
            except:
                pass
            
            # Limpiar estado si fue cancelado
            if self.proceso_cancelado:
                progreso_actual.update({
                    'estado': 'idle',
                    'mensaje': 'Proceso cancelado',
                    'inicio_proceso': None,
                    'thread_id': None
                })

    def guardar_resultados_en_excel(self):
        try:
            if not os.path.exists("reportes_excel"):
                os.makedirs("reportes_excel")
                
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            archivo = f"reportes_excel/reporte_simit_{timestamp}.xlsx"
            
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Control de Multas"

            # Colores y estilos
            verde_oscuro = "1F7246"
            verde_claro = "C6E0B4"
            rojo_claro = "FFE6E6"
            
            # Configuración de columnas
            ws1.column_dimensions['A'].width = 15
            ws1.column_dimensions['B'].width = 15
            ws1.column_dimensions['C'].width = 15
            ws1.column_dimensions['D'].width = 35
            ws1.column_dimensions['E'].width = 50

            # Título principal
            ws1.merge_cells('A1:E2')
            titulo = ws1.cell(row=1, column=1, value="FORMATO DE CONTROL DE MULTAS DE TRÁNSITO")
            titulo.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
            titulo.alignment = Alignment(horizontal="center", vertical="center")
            titulo.fill = PatternFill(start_color=verde_oscuro, end_color=verde_oscuro, fill_type="solid")

            # Fecha del reporte
            ws1.merge_cells('A3:E3')
            fecha = ws1.cell(row=3, column=1, value=f"Reporte generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            fecha.font = Font(name='Arial', size=10, italic=True)
            fecha.alignment = Alignment(horizontal="right")

            # Encabezados
            encabezados = ["Placa", "Estado Multas", "Resultado", "Evidencia", "Detalles"]
            for col, encabezado in enumerate(encabezados, 1):
                celda = ws1.cell(row=4, column=col, value=encabezado)
                celda.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
                celda.fill = PatternFill(start_color=verde_oscuro, end_color=verde_oscuro, fill_type="solid")
                celda.alignment = Alignment(horizontal="center", vertical="center")

            # Datos
            for idx, (placa, tiene_multa, resultado, captura, detalle_multas) in enumerate(self.resultados, 5):
                # Color de fila según estado
                if tiene_multa == "Sí":
                    fill_color = rojo_claro
                elif resultado == "Error":
                    fill_color = "FFCCCC"
                else:
                    fill_color = verde_claro if idx % 2 == 0 else "FFFFFF"
                
                datos_fila = [placa, tiene_multa, resultado, "Ver imagen adjunta", detalle_multas or "Sin detalles"]
                
                for col_idx, valor in enumerate(datos_fila, 1):
                    celda = ws1.cell(row=idx, column=col_idx, value=valor)
                    celda.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    if col_idx == 1:  # Placa
                        celda.alignment = Alignment(horizontal="center")
                    elif col_idx == 2 and tiene_multa == "Sí":  # Estado Multas
                        celda.font = Font(color="FF0000", bold=True)
                        celda.alignment = Alignment(horizontal="center")
                    elif col_idx == 5:  # Detalles
                        celda.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Agregar imagen
                if captura != "Sin captura" and os.path.exists(captura):
                    try:
                        img = Image(captura)
                        img.width = 300
                        img.height = 150
                        ws1.row_dimensions[idx].height = 120
                        ws1.add_image(img, f"D{idx}")
                    except Exception as e:
                        print(f"Error agregando imagen: {e}")

            # Guardar archivo
            wb.save(archivo)
            
            if os.path.exists(archivo) and os.path.getsize(archivo) > 1000:
                return archivo
            else:
                return None
            
        except Exception as e:
            print(f"Error generando Excel: {e}")
            return None

# RUTAS FLASK MEJORADAS
@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/iniciar_proceso', methods=['POST'])
def iniciar_proceso():
    global progreso_actual
    
    try:
        # Limpiar procesos colgados antes de iniciar
        limpiar_proceso_si_colgado()
        
        data = request.get_json()
        placas_texto = data.get('placas', '')
        placas = [placa.strip().upper() for placa in placas_texto.split('\n') if placa.strip()]
        
        if not placas:
            return jsonify({'error': 'No se ingresaron placas válidas'}), 400
        
        if progreso_actual['estado'] == 'processing':
            # Verificar si realmente está procesando o está colgado
            if (progreso_actual['inicio_proceso'] and 
                datetime.now() - progreso_actual['inicio_proceso'] > timedelta(seconds=30)):
                print("🧹 Forzando limpieza de proceso aparentemente colgado")
                limpiar_proceso_si_colgado()
            else:
                return jsonify({'error': 'Ya hay un proceso en ejecución. Espere unos minutos.'}), 400
        
        # Reiniciar progreso
        progreso_actual = {
            'estado': 'idle',
            'mensaje': 'Iniciando...',
            'placa_actual': '',
            'total': len(placas),
            'procesadas': 0,
            'porcentaje': 0,
            'resultados': [],
            'archivo_excel': '',
            'inicio_proceso': datetime.now(),
            'thread_id': None
        }
        
        scraper = SimitScraper()
        thread = threading.Thread(target=scraper.buscar_placas, args=(placas,))
        thread.daemon = True
        thread.start()
        
        progreso_actual['thread_id'] = thread.ident
        
        return jsonify({'success': True, 'mensaje': 'Proceso iniciado', 'total_placas': len(placas)})
        
    except Exception as e:
        print(f"Error en iniciar_proceso: {e}")
        return jsonify({'error': f'Error: {str(e)}'}), 500

@app.route('/progreso')
def obtener_progreso():
    global progreso_actual
    
    # Verificar si el proceso está colgado
    limpiar_proceso_si_colgado()
    
    return jsonify(progreso_actual.copy())

@app.route('/cancelar_proceso', methods=['POST'])
def cancelar_proceso():
    """Nueva ruta para cancelar procesos"""
    global progreso_actual
    
    if progreso_actual['estado'] == 'processing':
        print("🛑 Cancelando proceso...")
        limpiar_proceso_si_colgado()
        return jsonify({'success': True, 'mensaje': 'Proceso cancelado'})
    else:
        return jsonify({'error': 'No hay proceso activo'}), 400

@app.route('/descargar_excel')
def descargar_excel():
    global progreso_actual
    
    archivo_excel = progreso_actual.get('archivo_excel', '')
    
    if archivo_excel and os.path.exists(archivo_excel):
        try:
            return send_file(
                archivo_excel, 
                as_attachment=True,
                download_name=f"reporte_simit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            return jsonify({'error': f'Error enviando archivo: {str(e)}'}), 500
    else:
        return jsonify({'error': 'No hay archivo disponible'}), 404

# HTML TEMPLATE MEJORADO
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SIMIT Scraper</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        
        .container {
            max-width: 800px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            overflow: hidden;
        }
        
        .header {
            background: linear-gradient(135deg, #1F7246 0%, #2E8B57 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }
        
        .header h1 {
            font-size: 2.5em;
            margin-bottom: 10px;
        }
        
        .content {
            padding: 40px;
        }
        
        .input-group {
            margin-bottom: 30px;
        }
        
        .input-group label {
            display: block;
            font-weight: bold;
            margin-bottom: 10px;
            color: #333;
            font-size: 1.1em;
        }
        
        .placas-input {
            width: 100%;
            height: 200px;
            padding: 15px;
            border: 2px solid #ddd;
            border-radius: 10px;
            font-size: 16px;
            font-family: monospace;
            resize: vertical;
        }
        
        .placas-input:focus {
            outline: none;
            border-color: #1F7246;
        }
        
        .btn {
            background: linear-gradient(135deg, #1F7246 0%, #2E8B57 100%);
            color: white;
            border: none;
            padding: 15px 30px;
            font-size: 18px;
            border-radius: 10px;
            cursor: pointer;
            width: 100%;
            font-weight: bold;
            margin-bottom: 10px;
        }
        
        .btn:hover:not(:disabled) {
            transform: translateY(-2px);
        }
        
        .btn:disabled {
            background: #ccc;
            cursor: not-allowed;
            transform: none;
        }
        
        .btn-cancel {
            background: linear-gradient(135deg, #dc3545 0%, #c82333 100%);
            display: none;
        }
        
        .progress-container {
            display: none;
            margin-top: 30px;
            padding: 25px;
            background: #f8f9fa;
            border-radius: 15px;
            border-left: 5px solid #1F7246;
        }
        
        .progress-bar {
            width: 100%;
            height: 35px;
            background: #e9ecef;
            border-radius: 20px;
            overflow: hidden;
            margin: 20px 0;
            position: relative;
        }
        
        .progress-fill {
            height: 100%;
            background: linear-gradient(90deg, #1F7246, #2E8B57, #20c997);
            width: 0%;
            transition: width 0.5s ease-out;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: bold;
            font-size: 16px;
        }
        
        .progress-info {
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            gap: 15px;
            margin-top: 20px;
        }
        
        .progress-item {
            background: white;
            padding: 15px;
            border-radius: 10px;
            text-align: center;
        }
        
        .progress-item strong {
            display: block;
            color: #1F7246;
            font-size: 1.3em;
            margin-bottom: 5px;
        }
        
        .results-container {
            display: none;
            margin-top: 30px;
            padding: 25px;
            background: linear-gradient(135deg, #e8f5e8 0%, #d4edda 100%);
            border-radius: 15px;
            border: 2px solid #1F7246;
            text-align: center;
        }
        
        .download-btn {
            background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
            margin-top: 20px;
            padding: 15px 40px;
        }
        
        .status-message {
            padding: 15px;
            margin: 15px 0;
            border-radius: 10px;
            font-weight: bold;
        }
        
        .status-success {
            background: #d4edda;
            color: #155724;
            border: 2px solid #b8dabc;
        }
        
        .status-error {
            background: #f8d7da;
            color: #721c24;
            border: 2px solid #f1aeb5;
        }
        
        .status-info {
            background: #d1ecf1;
            color: #0c5460;
            border: 2px solid #abdde5;
        }
        
        .improvements-banner {
            background: linear-gradient(135deg, #17a2b8 0%, #138496 100%);
            color: white;
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 10px;
            text-align: center;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚗 SIMIT Scraper Optimizado</h1>
            <p>Sistema de Control de Multas - Versión Mejorada</p>
        </div>
        
        <div class="content">
            <div class="improvements-banner">
                <h4>🚀 Mejoras Implementadas</h4>
                <p>✅ Sin procesos colgados | ✅ Timeouts optimizados | ✅ Cancelación de procesos | ✅ Limpieza automática</p>
            </div>
            
            <div class="input-group">
                <label for="placas">Ingrese las placas (una por línea):</label>
                <textarea 
                    id="placas" 
                    class="placas-input" 
                    placeholder="ABC123&#10;DEF456&#10;GHI789"
                >ABC123
DEF456</textarea>
            </div>
            
            <button id="iniciarBtn" class="btn" onclick="iniciarProceso()">
                🚀 Iniciar Búsqueda
            </button>
            
            <button id="cancelarBtn" class="btn btn-cancel" onclick="cancelarProceso()">
                🛑 Cancelar Proceso
            </button>
            
            <div id="progressContainer" class="progress-container">
                <h3>Progreso del Proceso</h3>
                <div class="progress-bar">
                    <div id="progressFill" class="progress-fill">0%</div>
                </div>
                <div id="statusMessage" class="status-message status-info">
                    <span id="statusText">Iniciando proceso...</span>
                </div>
                <div class="progress-info">
                    <div class="progress-item">
                        <strong id="placaActual">-</strong>
                        <span>Placa Actual</span>
                    </div>
                    <div class="progress-item">
                        <strong id="contador">0 / 0</strong>
                        <span>Procesadas</span>
                    </div>
                    <div class="progress-item">
                        <strong id="estadoGeneral">Iniciando</strong>
                        <span>Estado</span>
                    </div>
                </div>
            </div>
            
            <div id="resultsContainer" class="results-container">
                <h3>🎉 ¡Proceso Completado!</h3>
                <p>El reporte Excel ha sido generado con todos los detalles de multas.</p>
                <button id="downloadBtn" class="btn download-btn" onclick="descargarExcel()">
                    📥 Descargar Reporte Excel
                </button>
            </div>
        </div>
    </div>

    <script>
        let intervalId = null;
        let procesoIniciado = false;
        let contadorErrores = 0;
        const MAX_ERRORES = 5;
        
        function iniciarProceso() {
            const placasTexto = document.getElementById('placas').value.trim();
            
            if (!placasTexto) {
                alert('Por favor, ingrese al menos una placa.');
                return;
            }
            
            const placasArray = placasTexto.split('\\n').filter(p => p.trim());
            
            if (placasArray.length === 0) {
                alert('No se encontraron placas válidas.');
                return;
            }
            
            if (!confirm(`¿Iniciar búsqueda para ${placasArray.length} placa(s)?`)) {
                return;
            }
            
            procesoIniciado = true;
            contadorErrores = 0;
            
            // Cambiar UI
            const btn = document.getElementById('iniciarBtn');
            const cancelBtn = document.getElementById('cancelarBtn');
            
            btn.disabled = true;
            btn.textContent = '⏳ Procesando...';
            cancelBtn.style.display = 'block';
            
            document.getElementById('progressContainer').style.display = 'block';
            document.getElementById('resultsContainer').style.display = 'none';
            
            // Resetear progreso
            const progressFill = document.getElementById('progressFill');
            progressFill.style.width = '0%';
            progressFill.textContent = '0%';
            
            // Iniciar proceso
            fetch('/iniciar_proceso', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json',
                },
                body: JSON.stringify({
                    placas: placasTexto
                })
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    intervalId = setInterval(actualizarProgreso, 1000);
                } else {
                    throw new Error(data.error || 'Error desconocido');
                }
            })
            .catch(error => {
                alert('Error al iniciar proceso: ' + error.message);
                resetearUI();
            });
        }
        
        function cancelarProceso() {
            if (!confirm('¿Está seguro que desea cancelar el proceso?')) {
                return;
            }
            
            fetch('/cancelar_proceso', {
                method: 'POST'
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    clearInterval(intervalId);
                    procesoIniciado = false;
                    resetearUI();
                    
                    const statusMessage = document.getElementById('statusMessage');
                    const statusText = document.getElementById('statusText');
                    statusMessage.className = 'status-message status-error';
                    statusText.textContent = 'Proceso cancelado por el usuario';
                }
            })
            .catch(error => {
                console.error('Error cancelando proceso:', error);
            });
        }
        
        function actualizarProgreso() {
            if (!procesoIniciado) return;
            
            fetch('/progreso')
            .then(response => response.json())
            .then(data => {
                contadorErrores = 0; // Reset error counter on success
                
                const porcentaje = Math.max(0, Math.min(100, data.porcentaje || 0));
                
                // Actualizar barra
                const progressFill = document.getElementById('progressFill');
                progressFill.style.width = porcentaje + '%';
                progressFill.textContent = porcentaje.toFixed(1) + '%';
                
                // Actualizar información
                document.getElementById('placaActual').textContent = data.placa_actual || '-';
                document.getElementById('contador').textContent = `${data.procesadas || 0} / ${data.total || 0}`;
                document.getElementById('estadoGeneral').textContent = data.estado || 'Procesando';
                
                // Actualizar mensaje
                const statusMessage = document.getElementById('statusMessage');
                const statusText = document.getElementById('statusText');
                statusText.textContent = data.mensaje || 'Procesando...';
                
                // Cambiar estilo según estado
                statusMessage.className = 'status-message';
                if (data.estado === 'completed') {
                    statusMessage.classList.add('status-success');
                    progressFill.style.width = '100%';
                    progressFill.textContent = '100%';
                    
                    clearInterval(intervalId);
                    procesoIniciado = false;
                    resetearUI();
                    
                    setTimeout(() => {
                        document.getElementById('resultsContainer').style.display = 'block';
                    }, 1000);
                    
                } else if (data.estado === 'error') {
                    statusMessage.classList.add('status-error');
                    clearInterval(intervalId);
                    procesoIniciado = false;
                    resetearUI();
                } else if (data.estado === 'idle' && procesoIniciado) {
                    // El proceso fue limpiado automáticamente
                    statusMessage.classList.add('status-error');
                    statusText.textContent = 'Proceso reiniciado automáticamente. Intente de nuevo.';
                    clearInterval(intervalId);
                    procesoIniciado = false;
                    resetearUI();
                } else {
                    statusMessage.classList.add('status-info');
                }
            })
            .catch(error => {
                contadorErrores++;
                console.error(`Error polling (${contadorErrores}/${MAX_ERRORES}):`, error);
                
                if (contadorErrores >= MAX_ERRORES) {
                    console.error('Muchos errores consecutivos, deteniendo polling');
                    clearInterval(intervalId);
                    procesoIniciado = false;
                    resetearUI();
                    
                    const statusMessage = document.getElementById('statusMessage');
                    const statusText = document.getElementById('statusText');
                    statusMessage.className = 'status-message status-error';
                    statusText.textContent = 'Conexión perdida. Refresque la página.';
                }
            });
        }
        
        function resetearUI() {
            const btn = document.getElementById('iniciarBtn');
            const cancelBtn = document.getElementById('cancelarBtn');
            
            btn.disabled = false;
            btn.textContent = '🚀 Iniciar Búsqueda';
            cancelBtn.style.display = 'none';
        }
        
        function descargarExcel() {
            const btn = document.getElementById('downloadBtn');
            btn.disabled = true;
            btn.textContent = '📥 Descargando...';
            
            fetch('/descargar_excel')
            .then(response => {
                if (!response.ok) {
                    throw new Error('Error en la descarga');
                }
                return response.blob();
            })
            .then(blob => {
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `reporte_simit_${new Date().toISOString().slice(0,10)}.xlsx`;
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
                window.URL.revokeObjectURL(url);
                
                btn.disabled = false;
                btn.textContent = '📥 Descargar Reporte Excel';
                
                alert('¡Archivo descargado exitosamente!');
            })
            .catch(error => {
                btn.disabled = false;
                btn.textContent = '📥 Descargar Reporte Excel';
                alert('Error al descargar: ' + error.message);
            });
        }
        
        // Auto-resize textarea
        document.getElementById('placas').addEventListener('input', function() {
            this.style.height = 'auto';
            this.style.height = Math.max(200, this.scrollHeight) + 'px';
        });
        
        // Prevenir cierre accidental durante proceso
        window.addEventListener('beforeunload', function(e) {
            if (procesoIniciado) {
                e.preventDefault();
                e.returnValue = '¿Estás seguro? Hay un proceso en ejecución.';
            }
        });
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
